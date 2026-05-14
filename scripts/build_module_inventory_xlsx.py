"""Build the FsNext module/UI inventory spreadsheet.

Output: docs/module_ui_inventory.xlsx — three sheets:
  1. "UI Inventory"     — flat table of every interactive element
  2. "Module Summary"   — per-module counts + status
  3. "Recommendations"  — analysis + recommended actions

Data sourced from manual scan of qml/**/*.qml (Aurora + Fshare variants),
captured 2026-05-13.  Re-run after any UI changes to refresh.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).resolve().parent.parent / "docs" / "module_ui_inventory.xlsx"

# ─────────────────────────────────────────────────────────────────────────
# Raw inventory.  Each row = (Module, Page, Section, ElementType, Label/ID, Action)
# ─────────────────────────────────────────────────────────────────────────
ROWS = [
    # ── Authentication ────────────────────────────────────────────────
    ("Authentication", "LoginView", "Form đăng nhập", "FsTextField", "emailField", "Nhập username/email"),
    ("Authentication", "LoginView", "Form đăng nhập", "FsTextField", "passwordField", "Nhập mật khẩu (ẩn ký tự), Enter để submit"),
    ("Authentication", "LoginView", "Tùy chọn", "MouseArea (checkbox)", "rememberMe", "Bật/tắt lưu thông tin đăng nhập"),
    ("Authentication", "LoginView", "Tùy chọn", "Text link", "Quên mật khẩu?", "(Chưa có hành động — placeholder)"),
    ("Authentication", "LoginView", "Tùy chọn", "Text link", "Đăng ký ngay →", "Mở trang đăng ký"),
    ("Authentication", "LoginView", "Button đăng nhập", "FsButton", "loginBtn", "Gọi authViewModel.login()"),
    ("Authentication", "LoginView", "Đăng nhập xã hội", "MouseArea (button)", "googleLogin", "Gọi authViewModel.loginWithGoogle()"),
    ("Authentication", "LoginView", "Đăng nhập xã hội", "MouseArea (button)", "facebookLogin", "Gọi authViewModel.loginWithFacebook()"),
    ("Authentication", "LoginView", "Đăng nhập xã hội", "MouseArea (button)", "zaloLogin", "Gọi authViewModel.loginWithFptId() (Zalo/FPT ID)"),
    ("Authentication", "Sidebar", "Logout", "MouseArea", "logoutBtn", "Gọi authViewModel.logout()"),

    # ── Navigation ────────────────────────────────────────────────────
    ("Navigation", "Main.qml", "Command Palette (Ctrl+K)", "FsCommandPalette", "cmdPalette", "Mở palette các lệnh điều hướng/hành động"),
    ("Navigation", "Main.qml", "Command Palette commands", "Command", "go.home / download / upload / sync / files / fav / account / settings", "Chuyển trang tương ứng"),
    ("Navigation", "Main.qml", "Command Palette commands", "Command", "act.add-dl", "Mở dialog thêm tải xuống"),
    ("Navigation", "Main.qml", "Command Palette commands", "Command", "act.logout", "Đăng xuất"),
    ("Navigation", "Main.qml", "Phím tắt toàn cục", "Shortcut", "Ctrl+K", "Bật/tắt command palette"),
    ("Navigation", "Main.qml", "Phím tắt toàn cục", "Shortcut", "Ctrl+Shift+D", "Toggle dark mode (dev build only)"),
    ("Navigation", "Main.qml", "Phím tắt toàn cục", "Shortcut", "Ctrl+V", "Dán link Fshare → chuyển sang Download"),
    ("Navigation", "Sidebar", "Hero card", "MouseArea", "upgradeButton", "Mở URL nâng cấp VIP"),
    ("Navigation", "Sidebar", "Menu chính", "NavItem", "Trang chủ", "currentPage = 8"),
    ("Navigation", "Sidebar", "Menu chính", "NavItem", "Tải xuống", "currentPage = 0"),
    ("Navigation", "Sidebar", "Menu chính", "NavItem", "Tải lên", "currentPage = 1"),
    ("Navigation", "Sidebar", "Menu chính", "NavItem", "Đồng bộ", "currentPage = 2"),
    ("Navigation", "Sidebar", "Menu chính", "NavItem", "My Files", "currentPage = 3"),
    ("Navigation", "Sidebar", "Menu chính", "NavItem", "Yêu thích", "currentPage = 4"),
    ("Navigation", "Sidebar", "Menu chính", "NavItem", "Tài khoản", "currentPage = 5"),
    ("Navigation", "Sidebar", "Menu chính", "NavItem", "Cài đặt", "currentPage = 6"),
    ("Navigation", "Sidebar", "Menu chính (dev)", "NavItem", "Showcase", "currentPage = 7 (dev build only)"),

    # ── Shell (global behaviors) ──────────────────────────────────────
    ("Shell", "Main.qml", "Drag-drop toàn cục", "DropArea", "globalDropArea", "Drop file → Upload; drop URL Fshare → Download"),
    ("Shell", "Main.qml", "Toast", "FsToast", "sessionToast", "Hiển thị thông báo phiên đăng nhập hết hạn"),

    # ── Dashboard (HomePage) ──────────────────────────────────────────
    ("Dashboard", "HomePage", "Thanh trên", "TextField", "searchInput", "Tìm file/folder hoặc dán link Fshare"),
    ("Dashboard", "HomePage", "Thanh trên", "MouseArea (icon)", "settingsButton", "Mở Settings (page 6)"),
    ("Dashboard", "HomePage", "Thanh trên", "MouseArea (button)", "uploadCTA", "Chuyển Upload (page 1)"),
    ("Dashboard", "HomePage", "Quick actions", "MouseArea (card)", "Dán link & tải", "Mở dialog thêm tải xuống"),
    ("Dashboard", "HomePage", "Quick actions", "MouseArea (card)", "Tải file lên", "Chuyển Upload (page 1)"),
    ("Dashboard", "HomePage", "Quick actions", "MouseArea (card)", "Thêm folder sync", "Chuyển Sync (page 2)"),
    ("Dashboard", "HomePage", "Quick actions", "MouseArea (card)", "Tạo link chia sẻ", "Chuyển File Manager (page 3)"),
    ("Dashboard", "HomePage", "Đang tải (continuation)", "MouseArea (button)", "pauseResumeButton", "Pause/resume transfer (per item)"),
    ("Dashboard", "HomePage", "Đang tải (continuation)", "MouseArea (button)", "openTransferButton", "Chuyển sang trang tương ứng"),
    ("Dashboard", "HomePage", "File gần đây - Filter", "MouseArea (tab)", "Tất cả / Video / Tài liệu / Ảnh", "Lọc danh sách file gần đây"),
    ("Dashboard", "HomePage", "File gần đây - Row", "MouseArea (icon)", "openFileButton", "Mở file với ứng dụng mặc định (per row)"),
    ("Dashboard", "HomePage", "File gần đây - Row", "MouseArea (icon)", "openFolderButton", "Mở thư mục chứa (per row)"),
    ("Dashboard", "HomePage", "File gần đây - Row", "MouseArea", "Row double-click", "Mở file (per row)"),

    # ── Download Management ───────────────────────────────────────────
    ("Download Management", "DownloadPage", "Header", "FsButton", "Pause/Resume all", "downloadViewModel.pauseAll() / resumeAll()"),
    ("Download Management", "DownloadPage", "Header", "FsButton", "Xoá lịch sử", "downloadViewModel.clearHistory()"),
    ("Download Management", "DownloadPage", "Header", "FsButton", "Thêm URL", "Mở addDialog"),
    ("Download Management", "DownloadPage", "Header", "FsSegmentedControl", "Đang tải / Lịch sử", "Chuyển tab active vs history"),
    ("Download Management", "DownloadPage", "Folder scan banner", "FsButton", "Huỷ scan", "downloadViewModel.cancelFolderScan()"),
    ("Download Management", "DownloadPage", "Add Download Dialog", "TextArea", "Links input", "Nhập một hoặc nhiều link Fshare"),
    ("Download Management", "DownloadPage", "Add Download Dialog", "FsFolderPicker", "Thư mục lưu", "Chọn thư mục đích trên máy"),
    ("Download Management", "DownloadPage", "Add Download Dialog", "FsTextField", "Mật khẩu (nếu có)", "Nhập mật khẩu file"),
    ("Download Management", "DownloadPage", "Add Download Dialog", "FsButton", "Huỷ / Tải xuống", "Đóng dialog hoặc downloadViewModel.addDownload()"),
    ("Download Management", "DownloadPage", "Transfer item row (per row)", "FsButton", "Pause", "downloadViewModel.pauseTask(id)"),
    ("Download Management", "DownloadPage", "Transfer item row (per row)", "FsButton", "Resume", "downloadViewModel.resumeTask(id)"),
    ("Download Management", "DownloadPage", "Transfer item row (per row)", "FsButton", "Cancel", "downloadViewModel.cancelTask(id)"),
    ("Download Management", "DownloadPage", "Transfer item row (per row)", "FsButton", "Dismiss (history)", "downloadViewModel.dismissCompleted(id)"),
    ("Download Management", "DownloadPage", "Transfer item row (per row)", "FsButton", "Open folder", "Mở thư mục chứa file đã tải"),

    # ── Upload Management ─────────────────────────────────────────────
    ("Upload Management", "UploadPage", "Header", "FsButton", "Pause/Resume all", "uploadViewModel.pauseAll() / resumeAll()"),
    ("Upload Management", "UploadPage", "Header", "FsButton", "Xoá lịch sử", "uploadViewModel.clearHistory()"),
    ("Upload Management", "UploadPage", "Header", "FsButton", "Thêm file", "Mở uploadDialog"),
    ("Upload Management", "UploadPage", "Header", "FsSegmentedControl", "Đang tải / Lịch sử", "Chuyển tab active vs history"),
    ("Upload Management", "UploadPage", "Drop zone", "Text (clickable)", "chọn từ máy", "Mở file picker"),
    ("Upload Management", "FsUploadDialog", "Drop area", "DropArea", "drop zone", "Drag file vào → thêm vào danh sách pending"),
    ("Upload Management", "FsUploadDialog", "Pending files list", "Text (clickable)", "chọn từ máy tính", "Mở _filePicker"),
    ("Upload Management", "FsUploadDialog", "Pending files list", "Text (clickable)", "Xoá tất cả", "Xoá toàn bộ pendingFiles"),
    ("Upload Management", "FsUploadDialog", "Pending file row", "FsButton", "X (per file)", "Xoá file đó khỏi pendingFiles"),
    ("Upload Management", "FsUploadDialog", "Tuỳ chọn", "FsSelect", "Tải vào", "Chọn thư mục đích trên Fshare"),
    ("Upload Management", "FsUploadDialog", "Tuỳ chọn", "FsTextField", "Mật khẩu bảo vệ", "Mật khẩu cho file upload"),
    ("Upload Management", "FsUploadDialog", "Tuỳ chọn", "FsSwitch", "File riêng tư", "Toggle private/secure flag"),
    ("Upload Management", "FsUploadDialog", "Footer", "FsButton", "Hủy / Bắt đầu tải lên", "Đóng dialog hoặc uploadViewModel.addUpload()"),
    ("Upload Management", "UploadPage", "Transfer item row (per row)", "FsButton", "Pause", "uploadViewModel.pauseTask(id)"),
    ("Upload Management", "UploadPage", "Transfer item row (per row)", "FsButton", "Resume", "uploadViewModel.resumeTask(id)"),
    ("Upload Management", "UploadPage", "Transfer item row (per row)", "FsButton", "Cancel", "uploadViewModel.cancelTask(id)"),
    ("Upload Management", "UploadPage", "Transfer item row (per row)", "FsButton", "Dismiss (history)", "uploadViewModel.dismissCompleted(id)"),
    ("Upload Management", "UploadPage", "Transfer item row (per row)", "FsButton", "Copy link", "uploadViewModel.copyLinkToClipboard(id)"),
    ("Upload Management", "UploadPage", "Transfer item row (per row)", "FsButton", "Open folder local", "uploadViewModel.revealLocalFile(id)"),
    ("Upload Management", "UploadPage", "Transfer item row (per row)", "FsButton", "Open in browser", "uploadViewModel.openShareLinkInBrowser(id)"),
    ("Upload Management", "UploadPage", "Transfer item row (per row)", "FsButton", "Show info", "Mở MyFiles tại folderId"),
    ("Upload Management", "UploadPage", "Transfer item row (per row)", "FsButton", "Delete local", "Confirm xoá file local sau upload"),
    ("Upload Management", "UploadPage", "Transfer item row (per row)", "FsButton", "Move up / down", "uploadViewModel.moveTaskUp() / moveTaskDown()"),

    # ── Sync (SyncPage) ───────────────────────────────────────────────
    ("Sync", "SyncPage", "Header", "FsButton", "Thêm folder sync", "Mở folderDlg"),
    ("Sync", "SyncPage", "Folder pair row", "MouseArea", "Row click", "syncViewModel.activeFolderId = folder.id"),
    ("Sync", "SyncPage", "Folder pair row (per row)", "FsButton", "Scan now", "Mở confirmScan dialog"),
    ("Sync", "SyncPage", "Folder pair row (per row)", "FsButton", "Open folder", "syncViewModel.openLocalFolder()"),
    ("Sync", "SyncPage", "Folder pair row (per row)", "FsButton", "Reset cache", "Mở confirmReset dialog"),
    ("Sync", "SyncPage", "Folder pair row (per row)", "FsButton", "Remove", "Mở confirmRemove dialog"),
    ("Sync", "SyncPage", "Active folder toolbar", "FsButton", "Đồng bộ ngay", "Trigger scan"),
    ("Sync", "SyncPage", "Active folder toolbar", "FsButton", "Mở thư mục", "Mở folder local"),
    ("Sync", "SyncPage", "Active folder toolbar", "FsButton", "Xoá cache", "Reset cache"),
    ("Sync", "SyncPage", "Active folder toolbar", "FsButton", "Xoá khỏi danh sách", "Remove folder pair"),
    ("Sync", "SyncPage", "Active folder toolbar", "FsSwitch", "Bật đồng bộ", "syncViewModel.setEnabled()"),
    ("Sync", "SyncPage", "Active folder toolbar", "FsSwitch", "Xoá local sau khi tải", "syncViewModel.setDeleteAfterUpload()"),
    ("Sync", "SyncPage", "File list (per row)", "FsButton", "Copy link", "syncViewModel.copyLinkFor()"),
    ("Sync", "SyncPage", "File list (per row)", "FsButton", "Open Fshare link", "syncViewModel.openFshareLink()"),
    ("Sync", "SyncPage", "Confirm dialogs", "FsButton", "Confirm scan / reset / remove", "Trigger action tương ứng"),
    ("Sync", "AddWatchFolderDialog", "Step 1", "FsFolderPicker", "Thư mục local", "Chọn folder local"),
    ("Sync", "AddWatchFolderDialog", "Step 1", "FsRemoteFolderTree", "Thư mục Fshare đích", "Chọn folder remote"),
    ("Sync", "AddWatchFolderDialog", "Step 2", "FsSwitch", "Theo dõi thư mục con", "watchSubfolders"),
    ("Sync", "AddWatchFolderDialog", "Step 2", "FsSwitch", "Xoá file local sau khi tải", "deleteAfterUpload"),
    ("Sync", "AddWatchFolderDialog", "Step 2", "FsTextField", "Bỏ qua file theo mẫu", "ignore patterns (e.g. *.tmp)"),
    ("Sync", "AddWatchFolderDialog", "Footer", "FsButton", "Huỷ / Quay lại / Tiếp tục / Thêm", "Navigate steps hoặc addFolder()"),
    ("Sync", "WatchFolderSettingsDialog", "Form", "FsSwitch", "Theo dõi thư mục con", "watchSubfolders"),
    ("Sync", "WatchFolderSettingsDialog", "Form", "FsSwitch", "Xoá file local sau khi tải", "deleteAfterUpload"),
    ("Sync", "WatchFolderSettingsDialog", "Form", "FsTextField", "Bỏ qua file theo mẫu", "ignore patterns"),
    ("Sync", "WatchFolderSettingsDialog", "Speed limit", "FsRadio", "Không giới hạn / Tùy chỉnh", "speedLimitMode = 0 hoặc 1"),
    ("Sync", "WatchFolderSettingsDialog", "Speed limit", "FsButton (spinner)", "+/-", "Tăng/giảm speedLimitKBps"),
    ("Sync", "WatchFolderSettingsDialog", "Footer", "FsButton", "Hủy / Lưu", "Đóng hoặc updateFolderSettings()"),
    ("Sync", "RemoveWatchDialog", "Footer", "FsButton", "Hủy / Xóa", "Đóng hoặc removeFolder()"),

    # ── Auto Upload (separate page, similar dialogs) ───────────────────
    ("Auto Upload", "AutoUploadPage", "Master toggle", "FsSwitch", "Bật đồng bộ tự động", "autoUploadViewModel.enabled"),
    ("Auto Upload", "AutoUploadPage", "Header", "FsButton", "Thêm thư mục", "Mở addDialog"),
    ("Auto Upload", "AutoUploadPage", "Section card", "FsButton", "Xóa lịch sử hoạt động", "autoUploadViewModel.clearActivity()"),
    ("Auto Upload", "AutoUploadPage", "Watch card (per card)", "FsButton", "Pause / Resume", "pauseFolder() / resumeFolder()"),
    ("Auto Upload", "AutoUploadPage", "Watch card (per card)", "FsButton", "Rescan", "rescanFolder()"),
    ("Auto Upload", "AutoUploadPage", "Watch card (per card)", "FsButton", "Retry", "retryFolder()"),
    ("Auto Upload", "AutoUploadPage", "Watch card (per card)", "FsButton", "Settings", "Mở settingsDialog"),
    ("Auto Upload", "AutoUploadPage", "Watch card (per card)", "FsButton", "Remove", "Mở removeDialog"),

    # ── Favorites ─────────────────────────────────────────────────────
    ("Favorites", "FavoritesPage", "Header", "TextInput", "Lọc theo đuôi file", "favoritesViewModel.extFilter"),
    ("Favorites", "FavoritesPage", "Header", "FsButton", "X (clear filter)", "Xoá nội dung filter input"),
    ("Favorites", "FavoritesPage", "Header", "FsButton", "Làm mới", "favoritesViewModel.loadFavorites()"),
    ("Favorites", "FavoritesPage", "Breadcrumb", "FsButton (icon)", "Heart (root)", "navigateToRoot()"),
    ("Favorites", "FavoritesPage", "Breadcrumb (per segment)", "Text (clickable)", "Tên folder", "navigateBack(segment)"),
    ("Favorites", "FavoritesPage", "File list", "MouseArea", "Left click row", "Toggle/single select"),
    ("Favorites", "FavoritesPage", "File list", "MouseArea", "Right click row", "Mở context menu"),
    ("Favorites", "FavoritesPage", "File list", "MouseArea", "Double click row", "Mở folder / play media / open file"),
    ("Favorites", "FavoritesPage", "File list (per row)", "FsButton", "Open in Explorer", "favoritesViewModel.openInExplorer()"),
    ("Favorites", "FavoritesPage", "Detail panel", "FsButton", "X (close)", "Đóng detail panel"),
    ("Favorites", "FavoritesPage", "Detail (single select)", "FsButton", "Mở thư mục / Xem trực tiếp / Mở file", "Navigate / play / open file đã tải"),
    ("Favorites", "FavoritesPage", "Detail (multi select)", "FsButton", "Sao chép link", "favoritesViewModel.copyLinks()"),
    ("Favorites", "FavoritesPage", "Detail", "FsButton", "Mở trên Fshare", "Mở URL fshare.vn"),
    ("Favorites", "FavoritesPage", "Detail", "FsButton", "Tải về", "downloadViewModel.addDownload()"),
    ("Favorites", "FavoritesPage", "Detail", "FsButton", "Bỏ yêu thích", "favoritesViewModel.removeFromFavorite()"),
    ("Favorites", "FavoritesPage", "Context menu", "MenuItem", "Open / Copy link / Remove / etc", "Các hành động cho file/folder yêu thích"),
    ("Favorites", "FavoritesPage", "Delete confirm dialog", "FsButton", "Hủy / Xóa", "Đóng hoặc deleteFiles()"),

    # ── User Account ──────────────────────────────────────────────────
    ("User Account", "UserInfoPage", "VIP card", "FsButton", "Refresh icon", "userInfoViewModel.refresh()"),
    ("User Account", "Sidebar", "Hero card", "MouseArea (button)", "Xem ưu đãi", "Mở https://www.fshare.vn/upgrade"),

    # ── Sharing ───────────────────────────────────────────────────────
    ("Sharing", "FsShareDialog", "Body", "FsButton", "Sao chép", "Copy shareUrl vào clipboard"),
    ("Sharing", "FsShareDialog", "Footer", "FsButton", "Đóng", "Đóng dialog"),

    # ── Settings ──────────────────────────────────────────────────────
    ("Settings", "SettingsPage", "Giao diện", "FsSwitch", "Tự động đăng nhập", "autoLogin"),
    ("Settings", "SettingsPage", "Giao diện", "FsSwitch", "Luôn hiển thị trên cùng", "stayOnTop"),
    ("Settings", "SettingsPage", "Giao diện", "FsSwitch", "Chế độ tối", "darkMode"),
    ("Settings", "SettingsPage", "Giao diện", "FsSwitch", "Thu nhỏ vào khay", "minimizeToTray"),
    ("Settings", "SettingsPage", "Khi tên file trùng", "FsRadio", "Đổi tên / Ghi đè / Bỏ qua / Hỏi mỗi lần", "duplicateNamePolicy (4 options)"),
    ("Settings", "SettingsPage", "Ngôn ngữ", "FsSegmentedControl", "Tiếng Việt / English", "languageViewModel.setLanguage()"),
    ("Settings", "SettingsPage", "Hiệu suất tải", "NumberSpinner", "Số download đồng thời", "downloadThreads (1–16)"),
    ("Settings", "SettingsPage", "Hiệu suất tải", "NumberSpinner", "Luồng song song/file", "downloadSegments (1–32)"),
    ("Settings", "SettingsPage", "Hiệu suất tải", "NumberSpinner", "Số upload đồng thời", "uploadThreads (1–8)"),
    ("Settings", "SettingsPage", "Hiệu suất tải", "NumberSpinner", "Tổng slot đồng thời", "maxGlobalSlots (0–32)"),
    ("Settings", "SettingsPage", "Tải xuống", "FsSwitch", "Tự động bắt đầu", "autoDownload"),
    ("Settings", "SettingsPage", "Tải xuống", "FsFolderPicker", "Thư mục tải xuống mặc định", "defaultDownloadDir"),
    ("Settings", "SettingsPage", "Kết nối", "FsSegmentedControl", "Proxy: Không / Hệ thống / Thủ công", "proxyMode (0/1/2)"),
    ("Settings", "SettingsPage", "Kết nối", "FsTextField", "Proxy host", "proxyHost"),
    ("Settings", "SettingsPage", "Kết nối", "NumberSpinner", "Proxy port", "proxyPort (1–65535)"),

    # ── File Management ───────────────────────────────────────────────
    ("File Management", "FileManagerPage", "Toolbar", "TextInput", "Tìm kiếm", "Search trong thư mục hiện tại"),
    ("File Management", "FileManagerPage", "Toolbar", "FsButton", "Thư mục mới", "Mở newFolderDialog"),
    ("File Management", "FileManagerPage", "Toolbar", "FsButton", "Làm mới", "fileManagerViewModel.refresh()"),
    ("File Management", "FileManagerPage", "Toolbar", "Toggle button", "List view / Grid view", "Chuyển chế độ hiển thị"),
    ("File Management", "FileManagerPage", "Breadcrumb", "FsButton", "Home", "Về root (My Files)"),
    ("File Management", "FileManagerPage", "Breadcrumb (per segment)", "Text (clickable)", "Tên folder", "Navigate to parent folder"),
    ("File Management", "FileManagerPage", "File list / Grid", "MouseArea", "Left click", "Chọn file/folder (single hoặc multi với Ctrl)"),
    ("File Management", "FileManagerPage", "File list / Grid", "MouseArea", "Double click", "Mở folder (navigate vào)"),
    ("File Management", "FileManagerPage", "File list / Grid", "MouseArea", "Right click", "Mở context menu"),
    ("File Management", "FileManagerPage", "File list / Grid (per row)", "FsButton (inline)", "Open in Explorer", "Mở file đã tải (nếu có)"),
    ("File Management", "FileManagerPage", "Detail panel", "FsButton", "X (close)", "Đóng detail panel"),
    ("File Management", "FileManagerPage", "Detail panel", "FsButton", "Mở thư mục / Phát / Xem trực tiếp", "Primary action theo loại file"),
    ("File Management", "FileManagerPage", "Detail panel", "FsButton", "Move / Copy", "Mở FileMoveCopyDialog"),
    ("File Management", "FileManagerPage", "Detail panel", "FsButton", "Copy link / Stream link", "Copy URL vào clipboard"),
    ("File Management", "FileManagerPage", "Detail panel", "FsButton", "Set password / Remove password", "Mở FilePasswordDialog"),
    ("File Management", "FileManagerPage", "Detail panel", "FsSwitch / Button", "Secure toggle", "Bật/tắt secure flag (single hoặc multi)"),
    ("File Management", "FileManagerPage", "Detail panel", "FsButton", "Share", "Sao chép link chia sẻ"),
    ("File Management", "FileManagerPage", "Detail panel", "FsButton", "Rename", "Mở FileRenameDialog"),
    ("File Management", "FileManagerPage", "Detail panel", "FsButton", "Open file / Open containing folder", "Mở local file đã tải"),
    ("File Management", "FileManagerPage", "Detail panel", "FsSwitch", "Direct link toggle", "Bật/tắt direct link"),
    ("File Management", "FileManagerPage", "Detail panel", "FsButton", "Delete", "Mở FileDeleteDialog"),
    ("File Management", "FileManagerPage", "Context menu (file)", "MenuItem", "Rename / Copy link / Open / Move / Copy", "Hành động file"),
    ("File Management", "FileManagerPage", "Context menu (file)", "MenuItem", "Secure / Remove secure / Set password", "Bảo mật"),
    ("File Management", "FileManagerPage", "Context menu (file)", "MenuItem", "Direct link / Play / Stream link / Delete", "Phát media + xoá"),
    ("File Management", "FileManagerPage", "Context menu (folder)", "MenuItem", "Rename / Copy link / Move / Copy / Open / Delete", "Hành động folder"),
    ("File Management", "FileRenameDialog", "Form", "FsTextField", "Tên mới", "Nhập tên mới"),
    ("File Management", "FileRenameDialog", "Footer", "FsButton", "Hủy / Đổi tên", "Đóng hoặc rename()"),
    ("File Management", "FileDeleteDialog", "Footer", "FsButton", "Hủy / Xóa", "Đóng hoặc delete()"),
    ("File Management", "FileMoveCopyDialog", "Tree", "MouseArea (tree node)", "Chọn folder đích", "Set target folder"),
    ("File Management", "FileMoveCopyDialog", "Tree", "MouseArea", "Root (My Files)", "Set target = root"),
    ("File Management", "FileMoveCopyDialog", "Footer", "FsButton", "Hủy / Move Here / Copy Here", "Đóng hoặc move()/copy()"),
    ("File Management", "FilePasswordDialog", "Form", "FsTextField", "Mật khẩu", "Nhập mật khẩu (hoặc trống để xoá)"),
    ("File Management", "FilePasswordDialog", "Form", "FsCheckbox", "Xoá mật khẩu hiện tại", "Toggle remove mode"),
    ("File Management", "FilePasswordDialog", "Footer", "FsButton", "Hủy / Set / Remove", "Đóng hoặc setPassword()/removePassword()"),
    ("File Management", "NewFolderDialog", "Form", "FsCheckbox", "Tạo bên trong folder đã chọn", "createInsideSelected"),
    ("File Management", "NewFolderDialog", "Form", "FsTextField", "Tên thư mục", "folderName"),
    ("File Management", "NewFolderDialog", "Footer", "FsButton", "Hủy / Tạo thư mục", "Đóng hoặc _submit()"),
]

# ─────────────────────────────────────────────────────────────────────────
# Module summary (curated, with status + description)
# ─────────────────────────────────────────────────────────────────────────
MODULES = [
    # name, primary pages, description, maturity, key risk
    ("Authentication", "LoginView, Sidebar", "Đăng nhập username/password + social (Google, Facebook, FPT ID), logout, remember me.", "Production", "Quên mật khẩu chưa có hành động — chỉ là placeholder link."),
    ("Navigation", "Sidebar, Main.qml (Command Palette, shortcuts)", "Sidebar 9 menu, Command Palette (Ctrl+K), global shortcuts (Ctrl+V để paste link, Ctrl+Shift+D toggle dark mode dev).", "Production", "Command Palette mới có 10 lệnh — có thể bổ sung nhiều action hơn (theme, language, recent files)."),
    ("Shell", "Main.qml", "Global drag-drop overlay, session-expired toast, page router, login/main shell switching.", "Production", "Drag-drop bị disable trước khi login — UX nên cho phép paste link ở login screen?"),
    ("Dashboard", "HomePage", "4 quick-action cards, search bar global, continuation transfer list, recent files filtered by type.", "Production", "Search là entry-point chính ở dashboard nhưng scope không rõ (global hay local). Quick-actions đều là shortcut tới page khác — có thể inline thêm tải xuống tại đây."),
    ("Download Management", "DownloadPage + FsAddDownloadDialog", "Tab Active/History, multi-link paste, password protected files, folder scan, per-row pause/resume/cancel.", "Production", "Add-download có FolderPicker nhưng không nhớ thư mục cuối cùng (cần check). Không có batch action 'Pause selected'."),
    ("Upload Management", "UploadPage + FsUploadDialog", "Drag-drop, folder picker target trên Fshare, password, secure flag, per-row 11 actions (move up/down, copy link, open browser…).", "Production", "Per-row có 11 nút inline — quá nhiều. Cần overflow menu (3-dot) cho actions ít dùng."),
    ("Sync", "SyncPage + AddWatchFolder/Settings/RemoveWatch dialogs", "Folder pair management, watch subfolders, ignore patterns, speed limit, scan now, reset cache.", "Production", "Duplicate với AutoUploadPage (xem khuyến nghị)."),
    ("Auto Upload", "AutoUploadPage + same dialogs as Sync", "Master toggle, per-folder pause/resume/rescan/retry/settings/remove cards.", "Unclear / Likely deprecated", "Trùng lặp hoàn toàn với Sync — chỉ khác UI layout. Cần xác định: legacy hay parallel?"),
    ("Favorites", "FavoritesPage", "Filter theo đuôi file, breadcrumb navigation, multi-select, detail panel với 8 hành động, context menu.", "Production", "Filter chỉ theo extension — chưa có filter theo size/date/tags."),
    ("User Account", "UserInfoPage + Sidebar hero", "VIP info, dung lượng, refresh button, link nâng cấp.", "Minimal", "Page rất sơ sài — chỉ refresh button. Thiếu: đổi mật khẩu, đăng xuất tất cả thiết bị, lịch sử đăng nhập, đổi email…"),
    ("Sharing", "FsShareDialog (mounted nhiều nơi)", "Copy link chia sẻ + đóng. Mounted từ FileManager, có thể từ Favorites.", "Minimal", "Không có advanced sharing: password riêng cho link, hết hạn, giới hạn lượt tải."),
    ("Settings", "SettingsPage", "Giao diện (4 switches), duplicate policy (4 radios), language (2 options), hiệu suất (4 spinners), tải xuống (switch + folder), kết nối/proxy (segmented + host + port).", "Production", "Không có save button — likely auto-save (cần verify). Thiếu: kích thước cache, lịch sử log, reset settings về mặc định, import/export."),
    ("File Management", "FileManagerPage + 7 sub-components/dialogs", "Folder tree + file list/grid + toolbar + breadcrumb + detail panel + 4 dialogs (Rename/Delete/Move-Copy/Password) + context menu file vs folder.", "Production", "Module phức tạp nhất (1880 dòng, vừa bị cắt cụt — đã fix). Không có drag-drop di chuyển file giữa folders trong UI (chỉ via dialog)."),
]

# ─────────────────────────────────────────────────────────────────────────
# Recommendations
# ─────────────────────────────────────────────────────────────────────────
RECS = [
    # priority, area, recommendation, rationale, effort
    ("Cao", "Sync vs Auto Upload trùng lặp",
     "Hợp nhất `Sync` và `Auto Upload` thành 1 module duy nhất; deprecate page còn lại (kiểm tra Main.qml currentPage 2 đang trỏ tới page nào).",
     "Hai page dùng cùng dialog (AddWatchFolderDialog, WatchFolderSettingsDialog, RemoveWatchDialog) và cùng ViewModel (autoUploadViewModel/syncViewModel — cần xác nhận liệu chúng có thực sự là 2 VM hay alias). Giữ cả hai gây nhầm lẫn cho user và double maintenance.",
     "Trung bình (1-2 ngày)"),

    ("Cao", "User Account quá sơ sài",
     "Mở rộng UserInfoPage: đổi mật khẩu, đăng xuất tất cả thiết bị, danh sách thiết bị đang đăng nhập, lịch sử transfer, đổi email/phone, kích hoạt 2FA.",
     "Hiện tại page chỉ có 1 nút refresh — không xứng với mức độ critical của module tài khoản. Đặc biệt 'đăng xuất tất cả thiết bị' là tính năng bảo mật cần thiết.",
     "Lớn (3-5 ngày tuỳ scope)"),

    ("Cao", "Forgot password chưa có hành động",
     "Implement flow quên mật khẩu (gọi API reset, hiển thị dialog nhập email/OTP, đặt mật khẩu mới).",
     "Hiện 'Quên mật khẩu?' chỉ là text link không click được — user bị mất tài khoản sẽ không có cách tự khôi phục.",
     "Trung bình (2-3 ngày + cần API backend)"),

    ("Trung bình", "Upload row có 11 nút inline — quá tải",
     "Gom các action ít dùng (Move up/down, Show info, Delete local, Open browser) vào 3-dot overflow menu. Giữ inline 4 nút chính: Pause/Resume/Cancel + Copy link.",
     "Mỗi row quá nhiều affordance làm khó tìm primary action và làm UI khó scan. Mobile-friendly principle: ≤5 nút inline.",
     "Nhỏ (½ ngày)"),

    ("Trung bình", "Sharing thiếu advanced options",
     "Mở rộng FsShareDialog: đặt mật khẩu cho link chia sẻ, ngày hết hạn, giới hạn lượt tải, thông kê lượt truy cập.",
     "Hiện chỉ có 'Sao chép URL' + 'Đóng' — thua sản phẩm cạnh tranh (Google Drive, OneDrive, Dropbox).",
     "Lớn (3-5 ngày, phụ thuộc API Fshare backend)"),

    ("Trung bình", "Settings thiếu reset & import/export",
     "Thêm nút 'Khôi phục mặc định' và 'Import/Export cấu hình' (JSON file).",
     "User reinstall hoặc move sang máy khác sẽ phải config lại tay. Settings có 15 fields, dễ sai lệch giữa máy.",
     "Nhỏ (1 ngày)"),

    ("Trung bình", "Command Palette mới có 10 lệnh",
     "Mở rộng: theme toggle, language switch, recent files, recent downloads, copy current URL, mở Settings tab N, focus search…",
     "Ctrl+K là power-user feature — càng nhiều shortcut càng giá trị. Hiện chủ yếu là nav (đã có sidebar).",
     "Nhỏ-Trung bình (1-2 ngày)"),

    ("Trung bình", "Filter Favorites quá hạn chế",
     "Thêm filter theo size (>=N MB), ngày upload, tag/folder, multi-extension (tick nhiều type).",
     "Hiện chỉ filter 1 extension at a time — user yêu thích >100 file sẽ khó tìm.",
     "Trung bình (2 ngày)"),

    ("Thấp", "Tray icon đã có nhưng menu tray ít action",
     "Thêm vào tray menu: Add download from clipboard, Recent uploads, Quick search, Tổng tốc độ.",
     "Tray là entry-point tiện lợi khi app minimize — chỉ Show/Pause/Quit là chưa khai thác hết.",
     "Nhỏ (½ ngày)"),

    ("Thấp", "Dashboard search ambiguous",
     "Clarify scope: thêm dropdown 'Search trong: [Tất cả file / Tải xuống / Tải lên / Yêu thích / Local]'. Hoặc dùng search syntax như Gmail.",
     "Hiện không rõ search field ở HomePage tra cứu gì — user paste link sẽ tự chuyển sang Download, nhưng query text thì sao?",
     "Nhỏ (½ ngày + cần thiết kế UX)"),

    ("Thấp", "File Manager không có drag-drop di chuyển file giữa folder",
     "Cho phép drag file/folder từ list/grid sang folder tree (sidebar) hoặc breadcrumb để move. Giữ Ctrl+drag = copy.",
     "Move/Copy hiện chỉ qua dialog — khó dùng khi cần move nhiều file. Drag-drop là chuẩn UX file manager.",
     "Trung bình (2 ngày + cần test edge cases)"),

    ("Thấp", "Không có module Notification trung tâm",
     "Thêm Notification Center (bell icon ở sidebar header): list các sự kiện gần đây — download done, upload done, sync error, VIP gần hết hạn, share link bị access nhiều bất thường.",
     "Hiện chỉ có 1 toast cho session-expired. User upload xong/download xong khi minimize sẽ không biết.",
     "Lớn (3-5 ngày)"),

    ("Thấp", "i18n incomplete",
     "Audit tất cả label hardcoded chưa qua qsTr(); chạy lupdate sinh .ts hoàn chỉnh; test toggle Vietnamese/English thực sự đổi mọi text.",
     "Đã có translation files (fshare_en.ts, fshare_vi.ts) nhưng chưa kiểm chứng 100% UI có dịch.",
     "Trung bình (2-3 ngày)"),

    ("Thấp", "Showcase page chỉ ở dev build",
     "Cân nhắc: gỡ Showcase khỏi production sidebar (hiện đã làm) nhưng giữ Ctrl+Shift+S để dev mở nhanh — đỡ phải build dev variant.",
     "Hiện logic isDevBuild flag — vẫn ổn, chỉ là gợi ý quality-of-life cho dev.",
     "Nhỏ (1 giờ)"),
]


# ─────────────────────────────────────────────────────────────────────────
# Build the workbook
# ─────────────────────────────────────────────────────────────────────────
wb = Workbook()

FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
SUB_FILL = PatternFill("solid", start_color="DCE6F1")
SUB_FONT = Font(name=FONT, bold=True, size=11)
BODY_FONT = Font(name=FONT, size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def style_header_row(sheet, row, end_col):
    for c in range(1, end_col + 1):
        cell = sheet.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def style_body(sheet, start_row, end_row, end_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, end_col + 1):
            cell = sheet.cell(row=r, column=c)
            if not cell.font or cell.font.name != FONT:
                cell.font = BODY_FONT
            cell.alignment = LEFT
            cell.border = BORDER


# ── Sheet 1: UI Inventory ──────────────────────────────────────────────
s1 = wb.active
s1.title = "UI Inventory"
HEADERS_1 = ["STT", "Module", "Page", "Section", "Element Type", "Element ID / Label", "Action / Purpose"]
s1.append(HEADERS_1)
style_header_row(s1, 1, len(HEADERS_1))

for i, (mod, page, section, etype, label, action) in enumerate(ROWS, start=1):
    s1.append([i, mod, page, section, etype, label, action])

style_body(s1, 2, len(ROWS) + 1, len(HEADERS_1))

widths_1 = [6, 22, 28, 32, 22, 38, 50]
for col_idx, w in enumerate(widths_1, start=1):
    s1.column_dimensions[get_column_letter(col_idx)].width = w
s1.row_dimensions[1].height = 28
s1.freeze_panes = "B2"

# Filter on data
s1.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS_1))}{len(ROWS) + 1}"

# ── Sheet 2: Module Summary ────────────────────────────────────────────
s2 = wb.create_sheet("Module Summary")
HEADERS_2 = ["STT", "Module", "Pages chính", "Mô tả", "Số element (tính từ Sheet 1)", "Mức độ hoàn thiện", "Rủi ro / vấn đề chính"]
s2.append(HEADERS_2)
style_header_row(s2, 1, len(HEADERS_2))

# Use COUNTIF formula referencing Sheet 1 column B (Module name)
inventory_module_range = f"'UI Inventory'!$B$2:$B${len(ROWS) + 1}"
for i, (name, pages, desc, maturity, risk) in enumerate(MODULES, start=1):
    excel_row = i + 1
    s2.append([
        i,
        name,
        pages,
        desc,
        f'=COUNTIF({inventory_module_range},B{excel_row})',
        maturity,
        risk,
    ])

style_body(s2, 2, len(MODULES) + 1, len(HEADERS_2))

# Totals row
total_row = len(MODULES) + 2
s2.cell(row=total_row, column=2, value="TỔNG").font = Font(name=FONT, bold=True, size=11)
s2.cell(row=total_row, column=5,
        value=f"=SUM(E2:E{len(MODULES) + 1})").font = Font(name=FONT, bold=True, size=11)
for c in range(1, len(HEADERS_2) + 1):
    cell = s2.cell(row=total_row, column=c)
    cell.fill = SUB_FILL
    cell.border = BORDER
    cell.alignment = LEFT if c not in (1, 5) else CENTER

widths_2 = [6, 22, 30, 50, 14, 18, 50]
for col_idx, w in enumerate(widths_2, start=1):
    s2.column_dimensions[get_column_letter(col_idx)].width = w
s2.row_dimensions[1].height = 28
s2.freeze_panes = "B2"

# ── Sheet 3: Recommendations ───────────────────────────────────────────
s3 = wb.create_sheet("Recommendations")
HEADERS_3 = ["STT", "Mức ưu tiên", "Khu vực / Module", "Khuyến nghị", "Lý do", "Ước tính công sức"]
s3.append(HEADERS_3)
style_header_row(s3, 1, len(HEADERS_3))

PRIORITY_COLOR = {"Cao": "F8CBAD", "Trung bình": "FFE699", "Thấp": "C6EFCE"}

for i, (prio, area, rec, why, effort) in enumerate(RECS, start=1):
    row = i + 1
    s3.append([i, prio, area, rec, why, effort])
    # Color the priority cell
    pc = s3.cell(row=row, column=2)
    pc.fill = PatternFill("solid", start_color=PRIORITY_COLOR.get(prio, "FFFFFF"))
    pc.font = Font(name=FONT, bold=True, size=10)

style_body(s3, 2, len(RECS) + 1, len(HEADERS_3))

widths_3 = [6, 14, 28, 60, 60, 22]
for col_idx, w in enumerate(widths_3, start=1):
    s3.column_dimensions[get_column_letter(col_idx)].width = w
s3.row_dimensions[1].height = 28
s3.freeze_panes = "B2"

# ── Save ───────────────────────────────────────────────────────────────
OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"Wrote {OUT}")
print(f"  Sheet 'UI Inventory': {len(ROWS)} rows")
print(f"  Sheet 'Module Summary': {len(MODULES)} modules")
print(f"  Sheet 'Recommendations': {len(RECS)} recommendations")
